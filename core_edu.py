"""
core_edu.py — Buscador de Recursos Educativos MINEDU
=====================================================
Flujo:
  - Indexacion manual: subida de PDF + metadata manual
  - Indexacion por Excel MINEDU: descarga paralela + vectorizacion
  - Busqueda con filtros (tipo_recurso, nivel, area, categoria)
  - Busqueda por pagina similar
"""

import json
import logging
import os
import re
import time
import asyncio
import aiohttp
import tempfile
from pathlib import Path
from typing import Optional, AsyncGenerator
from io import BytesIO

from PIL import Image

import core

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuracion
# ---------------------------------------------------------------------------
EDU_COLLECTION   = os.getenv("EDU_COLLECTION",    "imagenes_embeddings")
EDU_INDEX_FILE   = os.getenv("EDU_INDEX_FILE",    "edu_index.json")
EDU_SEARCH_LIMIT = int(os.getenv("EDU_SEARCH_LIMIT", "5"))

# Columnas del Excel MINEDU que se guardan como metadata en Qdrant
EXCEL_META_FIELDS = [
    "categoria", "sub_categoria", "tipo_recurso", "titulo", "modalidad", "servicio_educativo",
    "autor", "derecho_autoridad", "anio_edicion", "lengua_idioma",
    "nivel", "area", "resumen","competencias",
]

# Mapeo nombre columna Excel -> clave interna
EXCEL_COL_MAP = {
    "CATEGORÍA":          "categoria",
    "SUB CATEGORÍA":      "sub_categoria",
    "2° SUB CATEGORÍA":   "sub_categoria_2",
    "3° SUB CATEGORÍA":   "sub_categoria_3",
    "4° SUB CATEGORÍA":   "sub_categoria_4",
    "TIPO RECURSO":       "tipo_recurso",
    "TÍTULO":             "titulo",
    "MODALIDAD":          "modalidad",
    "SERVICIO EDUCATIVO": "servicio_educativo",
    "RUTA BANNER":        "ruta_banner",
    "AUTOR":              "autor",
    "DERECHO AUTORIDAD":  "derecho_autoridad",
    "AÑO EDICIÓN":        "anio_edicion",
    "LENGUA/IDIOMA":      "lengua_idioma",
    "NIVEL":              "nivel",
    "ÁREA":               "area",
    "CICLO":              "ciclo",
    "EDAD Ó GRADO":       "edad_o_grado",
    "FECHA PUBLICACIÓN":  "fecha_publicacion",
    "TIPO ENLACE":        "tipo_enlace",
    "ENLACE":             "enlace",
    "RESUMEN":            "resumen",
    "ENFOQUES":           "enfoques",
    "COMPETENCIAS":       "competencias",
    "ESTADO":             "estado",
    "ESTRATEGIA":         "estrategia",
    "ID":                 "id_excel",
}

# ---------------------------------------------------------------------------
# Indice local JSON
# ---------------------------------------------------------------------------

def _load_index() -> dict:
    p = Path(EDU_INDEX_FILE)
    if p.exists():
        with open(p, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"resources": {}}


def _save_index(index: dict) -> None:
    with open(EDU_INDEX_FILE, "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)





# ---------------------------------------------------------------------------
# Validacion de documentos ya vectorizados
# ---------------------------------------------------------------------------

# Flag global para cancelar procesamiento
_processing_cancelled = False

def cancel_processing():
    """Cancela el procesamiento en curso del Excel."""
    global _processing_cancelled
    _processing_cancelled = True

def _normalize_title(title: str) -> str:
    """Normaliza un título para comparación (minúsculas, sin espacios extras, sin caracteres especiales)."""
    import unicodedata
    # Convertir a minúsculas y quitar acentos
    nfd = unicodedata.normalize('NFD', title.lower())
    title_clean = ''.join(c for c in nfd if unicodedata.category(c) != 'Mn')
    # Quitar caracteres especiales y espacios extras
    title_clean = re.sub(r'[^a-z0-9\s]', '', title_clean)
    title_clean = re.sub(r'\s+', ' ', title_clean).strip()
    return title_clean

def _is_document_already_indexed(filename: str, titulo: str = None) -> tuple[bool, dict]:
    """
    Verifica si un documento ya está indexado en la base de datos.
    Compara por nombre de archivo Y opcionalmente por título normalizado.
    
    Retorna: (is_indexed: bool, document_info: dict)
      - is_indexed: True si el documento ya fue procesado
      - document_info: diccionario con la info del documento (indexed_at, pages, titulo, etc.)
    """
    index = _load_index()
    
    # Buscar por nombre exacto primero
    if filename in index["resources"]:
        doc_info = index["resources"][filename]
        if doc_info.get("indexed"):
            return True, doc_info
    
    # Si se proporciona título, buscar por título normalizado (más robusto)
    if titulo:
        titulo_norm = _normalize_title(titulo)
        for fname, doc_info in index["resources"].items():
            if doc_info.get("indexed"):
                existing_titulo = doc_info.get("titulo", "")
                if _normalize_title(existing_titulo) == titulo_norm:
                    return True, doc_info
    
    return False, {}


# ---------------------------------------------------------------------------
# Escaneo de carpeta local
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Indexacion manual (un PDF + metadata manual)
# ---------------------------------------------------------------------------

def index_manual_upload(pdf_bytes: bytes, filename: str, metadata: dict) -> dict:
    """
    Indexa un PDF subido manualmente con metadata proporcionada por el usuario.
    metadata puede contener: titulo, tipo_recurso, nivel, area, categoria, autor, etc.
    
    Valida primero si el documento ya está indexado.
    Si ya existe, retorna un error sin procesar.
    """
    # --- Validacion: ¿Ya existe indexado? ---
    titulo = metadata.get("titulo", filename)
    is_already_indexed, existing_info = _is_document_already_indexed(filename, titulo)
    if is_already_indexed:
        indexed_date = existing_info.get("indexed_at", "desconocida")
        pages = existing_info.get("pages", 0)
        existing_titulo = existing_info.get("titulo", filename)
        error_msg = (
            f"⚠️ DOCUMENTO YA INDEXADO\n"
            f"Nombre: {filename}\n"
            f"Título: {existing_titulo}\n"
            f"Fecha indexación: {indexed_date}\n"
            f"Páginas procesadas: {pages}\n\n"
            f"Este documento ya se encuentra en la base de datos y no puede ser reprocesado."
        )
        return {
            "error": True,
            "status": "skipped",
            "message": error_msg,
            "filename": filename,
            "reason": "Document already indexed"
        }
    
    # --- Procesar documento nuevo ---
    index = _load_index()
    index["resources"][filename] = {
        "path": None, "indexed": False, "indexed_at": None, "pages": 0,
        **{k: v for k, v in metadata.items() if v},
    }

    result = core.index_pdf(
        pdf_bytes=pdf_bytes,
        collection_name=EDU_COLLECTION,
        filename=filename,
        extra_payload={k: v for k, v in metadata.items() if v},
    )

    index["resources"][filename]["indexed"]    = True
    index["resources"][filename]["indexed_at"] = time.strftime("%Y-%m-%dT%H:%M:%S")
    index["resources"][filename]["pages"]      = result["pages_indexed"]
    _save_index(index)

    return {**result, **metadata}


# ---------------------------------------------------------------------------
# Lectura del Excel MINEDU
# ---------------------------------------------------------------------------

def parse_excel_minedu(excel_path: str) -> list[dict]:
    """
    Lee el Excel MINEDU y retorna solo las filas validas:
      - LENGUA/IDIOMA contiene 'Castellano'
      - TIPO ENLACE == 'PDF'
      - ESTADO == 'Activo'
      - ENLACE no nulo

    Retorna lista de dicts con claves internas (EXCEL_COL_MAP).
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("Instala openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    # Leer headers de la primera fila
    raw_headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {h: i for i, h in enumerate(raw_headers) if h}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def get(col_name):
            idx = col_idx.get(col_name)
            if idx is None:
                return None
            val = row[idx]
            return str(val).strip() if val is not None else None

        # Filtros obligatorios
        lengua    = get("LENGUA/IDIOMA") or ""
        tipo_enl  = get("TIPO ENLACE")   or ""
        estado    = get("ESTADO")        or ""
        enlace    = get("ENLACE")

        if "castellano" not in lengua.lower():  continue
        if tipo_enl.upper() != "PDF":           continue
        if estado.lower() != "activo":          continue
        if not enlace:                          continue

        # Construir metadata con claves internas
        entry = {}
        for excel_col, internal_key in EXCEL_COL_MAP.items():
            val = get(excel_col)
            entry[internal_key] = val if val and val != "--" else None

        rows.append(entry)

    logger.info("Excel MINEDU: %d filas validas de %d totales", len(rows), ws.max_row - 1)
    return rows


# ---------------------------------------------------------------------------
# Descarga paralela + indexacion desde Excel (con progreso via SSE)
# ---------------------------------------------------------------------------

async def _download_pdf(session: aiohttp.ClientSession, url: str, dest: Path) -> bool:
    """Descarga un PDF desde una URL. Retorna True si exitoso."""
    try:
        async with session.get(url, timeout=aiohttp.ClientTimeout(total=60)) as resp:
            if resp.status != 200:
                return False
            content = await resp.read()
            dest.write_bytes(content)
            return True
    except Exception as e:
        logger.warning("Error descargando %s: %s", url, e)
        return False


async def process_excel_minedu(excel_path: str, max_concurrent: int = 4) -> AsyncGenerator[dict, None]:
    """
    Pipeline completo desde Excel MINEDU:
      1. Lee y filtra filas validas
      2. Descarga PDFs en paralelo (max_concurrent a la vez)
      3. Vectoriza e inserta en Qdrant con metadata completa
      4. Emite eventos de progreso como dicts

    Uso: async for event in process_excel_minedu(path): ...
    """
    rows = parse_excel_minedu(excel_path)
    total = len(rows)

    if total == 0:
        yield {"type": "done", "total": 0, "success": 0, "failed": 0,
               "msg": "No se encontraron filas validas (Castellano + PDF + Activo)."}
        return

    yield {"type": "start", "total": total, "msg": f"{total} recursos validos encontrados en el Excel"}

    index = _load_index()
    temp_dir = Path(tempfile.gettempdir())
    dest_base = temp_dir / "minedu_downloads"
    dest_base.mkdir(exist_ok=True)

    success_count = 0
    failed_count  = 0
    semaphore     = asyncio.Semaphore(max_concurrent)

    async def process_one(row: dict, pos: int) -> dict:
        """Descarga + vectoriza un recurso. Retorna evento de resultado."""
        global _processing_cancelled
        
        # Verificar si fue cancelado
        if _processing_cancelled:
            titulo = row.get("titulo", "sin_titulo")
            return {
                "type": "cancelled",
                "pos": pos,
                "total": total,
                "msg": f"[{pos}/{total}] ⏸️ CANCELADO: procesamiento detenido por el usuario"
            }
        
        titulo  = row.get("titulo") or row.get("enlace", "sin_titulo")
        enlace  = row.get("enlace", "")
        # Nombre de archivo: sanitizado desde titulo o URL
        safe_name = re.sub(r'[^a-zA-Z0-9_\-]', '_', titulo[:60]) + ".pdf"
        temp_dir = Path(tempfile.gettempdir())
        dest_base = temp_dir / "minedu_downloads"
        dest_base.mkdir(exist_ok=True)

        # --- Validacion: ¿Ya existe indexado? (por nombre Y título normalizado) ---
        is_already_indexed, existing_info = _is_document_already_indexed(safe_name, titulo)
        if is_already_indexed:
            indexed_date = existing_info.get("indexed_at", "desconocida")
            pages = existing_info.get("pages", 0)
            return {
                "type": "skipped",
                "pos": pos,
                "total": total,
                "filename": safe_name,
                "titulo": titulo,
                "indexed_at": indexed_date,
                "pages": pages,
                "msg": f"[{pos}/{total}] ⏭️  OMITIDO (ya indexado): {titulo[:50]} — Procesado: {indexed_date}"
            }

        async with semaphore:
            # --- Descarga ---
            dest = dest_base / safe_name
            async with aiohttp.ClientSession() as session:
                ok = await _download_pdf(session, enlace, dest)

            if not ok:
                return {"type": "error", "pos": pos, "total": total,
                        "filename": safe_name, "titulo": titulo,
                        "msg": f"[{pos}/{total}] Error descargando: {enlace}"}

            # --- Vectorizacion ---
            try:
                pdf_bytes = dest.read_bytes()
                # Construir payload con toda la metadata del Excel
                payload = {k: v for k, v in row.items() if v is not None and k != "id_excel"}

                result = core.index_pdf(
                    pdf_bytes=pdf_bytes,
                    collection_name=EDU_COLLECTION,
                    filename=safe_name,
                    extra_payload=payload,
                )

                # Registrar en indice local
                index["resources"][safe_name] = {
                    "path":       str(dest.resolve()),
                    "indexed":    True,
                    "indexed_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
                    "pages":      result["pages_indexed"],
                    **payload,
                }

                return {"type": "ok", "pos": pos, "total": total,
                        "filename": safe_name, "titulo": titulo,
                        "pages": result["pages_indexed"],
                        "msg": f"[{pos}/{total}] ✓ {titulo[:50]} — {result['pages_indexed']} págs."}

            except Exception as e:
                return {"type": "error", "pos": pos, "total": total,
                        "filename": safe_name, "titulo": titulo,
                        "msg": f"[{pos}/{total}] Error vectorizando {safe_name}: {e}"}

    # Procesar en paralelo con semaforo
    tasks = [process_one(row, i + 1) for i, row in enumerate(rows)]

    skipped_count = 0
    cancelled = False
    for coro in asyncio.as_completed(tasks):
        event = await coro
        if event["type"] == "ok":
            success_count += 1
        elif event["type"] == "skipped":
            skipped_count += 1
        elif event["type"] == "cancelled":
            cancelled = True
            break  # Salir del loop si fue cancelado
        else:
            failed_count += 1
        yield event

    # Guardar indice al final
    _save_index(index)

    # Resetear flag de cancelación
    global _processing_cancelled
    _processing_cancelled = False

    yield {
        "type":      "done",
        "total":     total,
        "success":   success_count,
        "failed":    failed_count,
        "skipped":   skipped_count,
        "cancelled": cancelled,
        "msg":       f"Completado: {success_count} indexados, {skipped_count} omitidos (ya procesados), {failed_count} fallidos." if not cancelled 
                     else f"Cancelado: {success_count} indexados, {skipped_count} omitidos, {failed_count} fallidos de {total} recursos.",
    }


# ---------------------------------------------------------------------------
# Busqueda con filtros
# ---------------------------------------------------------------------------

def _build_filter(tipo_recurso=None, nivel=None, area=None, categoria=None):
    from qdrant_client.http import models as qm
    conditions = []
    if tipo_recurso:
        conditions.append(qm.FieldCondition(key="tipo_recurso", match=qm.MatchValue(value=tipo_recurso)))
    if nivel:
        conditions.append(qm.FieldCondition(key="nivel",        match=qm.MatchValue(value=nivel)))
    if area:
        conditions.append(qm.FieldCondition(key="area",         match=qm.MatchValue(value=area)))
    if categoria:
        conditions.append(qm.FieldCondition(key="categoria",    match=qm.MatchValue(value=categoria)))
    return qm.Filter(must=conditions) if conditions else None


def _run_search_edu(multivector, limit, tipo_recurso=None, nivel=None, area=None, categoria=None) -> list[dict]:
    from qdrant_client.http import models as qm
    client = core.get_qdrant()
    t0 = time.time()

    result = client.query_points(
        collection_name=EDU_COLLECTION,
        query=multivector,
        limit=limit,
        query_filter=_build_filter(tipo_recurso, nivel, area, categoria),
        timeout=100,
        search_params=qm.SearchParams(
            quantization=qm.QuantizationSearchParams(ignore=False, rescore=True, oversampling=2.0)
        ),
    )

    logger.info("Busqueda educativa completada en %.4fs — %d resultados", time.time() - t0, len(result.points))

    hits = []
    for point in result.points:
        p = point.payload or {}
        hits.append({
            "id":           point.id,
            "score":        point.score,
            "image_base64": p.get("image_base64"),
            "image_name":   p.get("image_name"),
            "source_file":  p.get("source_file"),
            "page_number":  p.get("page_number"),
            "total_pages":  p.get("total_pages"),
            # Metadata para filtros
            "tipo_recurso": p.get("tipo_recurso"),
            "nivel":        p.get("nivel"),
            "area":         p.get("area"),
            "categoria":    p.get("categoria"),
            # Metadata de detalle
            "titulo":           p.get("titulo"),
            "modalidad":        p.get("modalidad"),
            "servicio_educativo": p.get("servicio_educativo"),
            "autor":            p.get("autor"),
            "derecho_autoridad": p.get("derecho_autoridad"),
            "anio_edicion":     p.get("anio_edicion"),
            "lengua_idioma":    p.get("lengua_idioma"),
            "resumen":          p.get("resumen"),
            "competencias":     p.get("competencias"),
        })
    return hits


def search_by_text(query, limit=EDU_SEARCH_LIMIT, tipo_recurso=None, nivel=None, area=None, categoria=None):
    logger.info("Busqueda por texto: '%s'", query)
    return _run_search_edu(core.embed_query_text(query), limit, tipo_recurso, nivel, area, categoria)


def search_by_image(image: Image.Image, limit=EDU_SEARCH_LIMIT, tipo_recurso=None, nivel=None, area=None, categoria=None):
    logger.info("Busqueda por imagen")
    return _run_search_edu(core.embed_query_image(image), limit, tipo_recurso, nivel, area, categoria)


def search_similar_to_page(image_b64: str, limit=EDU_SEARCH_LIMIT) -> list[dict]:
    """Busca paginas visualmente similares dado un base64 de pagina."""
    import base64
    logger.info("Busqueda por pagina similar")
    image = Image.open(BytesIO(base64.b64decode(image_b64))).convert("RGB")
    return _run_search_edu(core.embed_query_image(image), limit)
